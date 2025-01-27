"""Importing"""
import datetime
import openpyxl
import glob
from openpyxl import load_workbook
import excel_editor


class Client:
    def __init__(self, name):
        self.name = name

    def __str__(self):
        return f"{self.name}"

    def client_name(self):
        name = self.split()
        return name

class ClientCar:
    def __init__(self, brand, model, year, engine, vin):
        self.brand = brand
        self.model = model
        self.year = year
        self.engine = engine
        self.vin = vin

def main():
    """Check if new or old client and handle it."""
    old_or_new = old_or_new_client_func()
    if old_or_new is True:
        excel_editor.main_excel()
    elif old_or_new is False:
        old_client_file_path = getting_old_client_file_func()
        old_client_new_repair_func(old_client_file_path)
        which_column_func(old_client_file_path)


"""Check if new or old client()"""
def old_or_new_client_func():
    new_column = input("Czy nowy klient? (y/n)").lower()
    if new_column in  ["n", "nie", "not"]:
        return False
    else:
        return True

def new_client_func():
    pass

"""If old client, get old file report and return path to file"""
def getting_old_client_file_func():
    client_path = glob.glob('**/*.xlsx', recursive = True)
    print(f"Wybierz klienta z listy:")
    order = 1
    book = {}
    for file in client_path:
        print(f"{order}.{file}")
        book.update({order: file})
        order += 1
    chosen_client_file = int(input(f"Wpisz numer klienta z listy: "))
    for ordered_numb, client_dir in book.items():
        if ordered_numb == chosen_client_file:
            print(f"client: {client_dir}")
            return client_dir

"""Insert informations to report"""
def old_client_new_repair_func(client_dir):
    old_client = load_workbook(client_dir, read_only=False)
    # Open first sheet
    f_sheet = old_client.active
    # Insert data in proper columns/tables
    f_sheet["H6"] = "Hello test"
    # Save document as update file
    old_file_name = f"{client_dir}"
    old_client.save(filename=f"{old_file_name}")

"""Handle which column is free and fill it up"""
def which_column_func(client_dir):
    columns = dict(C=" ", D=" ", E=" ", F=" ", G=" ", H=" ")
    wb = load_workbook(client_dir, data_only=True)
    f_sheet = wb.active
    start_col = 2  # 'C' column index
    end_col = 7  # 'H' column index
    """Read all values from cells from row and add to list"""
    for i in range(14, 15):
        row = [cell.value for cell in f_sheet[i][start_col:end_col + 1]]
    """Add to dictionary value to specified column"""
    counter = 0
    for key, value in columns.items():
        columns.update({key:row[counter]})
        counter += 1
    free_column_func(columns)

"""Check columns for free space and return first free column to use"""
def free_column_func(columns):
    col_letter_list = list()
    for key, value in columns.items():
        if value is None:
            col_letter_list += key
    return col_letter_list[0]

        #TODO transfer the free column letter from fee_column_func to excel editor and change columns to write
        #TODO next things to fullfill next columns, check which column is free and use it.
        #TODO Make new column func and main with openpyxl in another smallers functions/classes - no big code blocks!


def client_name_func():
    """Obtaining client's detail"""
    client_name = input("Imię i nazwisko klienta: ")
    return Client.client_name(client_name)

def client_car_func():
    """Obtaining vehicle's detail"""
    while True:
        car_param = {"Marka":"brak", "Model":"brak", "Rok":"brak", "Silnik":"brak", "Numer rejestracji":"brak", "VIN":"brak", "Przebieg":"brak"}
        for key in car_param:
            value = input(f"Podaj {key}: ")
            if value:
                car_param.update({key: value})
            else:
                print(f"Nie wpisałeś: {key}")
                continue
        break
    return car_param

def dates_func():
    """Obtaining date and checking if proper format"""
    while True:
        try:
            date_format = "%d.%m.%Y"
            accept_date = input("Data przyjęcia pojazdu DD.MM.RRRR: ")
            accept_date = datetime.datetime.strptime(accept_date, date_format)
            end_date = input("Data wydania pojazdu DD.MM.RRRR: ")
            end_date = datetime.datetime.strptime(end_date, date_format)
        except ValueError:
            print("Niepoprawny format daty, prawidłowy format: DD.MM.RRRR")
            continue
        break
    return accept_date, end_date

def checklist_func():
    """Checklist after car verification"""
    while True:
        car_checklist = {"Zawieszenie":"brak", "Oświetlenie":"brak", "Klimatyzacja":"brak",
                         "Silnik":"brak", "Koła":"brak", "Hamulce":"brak", "Nadwozie":"brak",
                         "Podwozie":"brak", "Korozja":"brak",}
        for key in car_checklist:
            value = input(f"Checklista {key}: ")
            if value:
                car_checklist.update({key: value})
            else:
                print(f"Nie wpisałeś: {key}")
                continue
        break
    return car_checklist

def todo_func():
    """Inputing problems with car and asking if there is more"""
    todo_list = []
    while True:
        customer_report = input("Awaria, naprawa zgłoszona przez klienta: ")
        todo_list.append(customer_report)
        next_report = input("Chcesz dodać kolejną usterkę? y/n ").lower()
        if next_report == "y":
            continue
        elif next_report != "n":
            continue
        else:
            break
    return todo_list

def repaired_func():
    """Inputing reapired items and asking if more"""
    repaired_list = {}
    while True:
        repaired_thing = input("Jaka naprawa została wykonana?: ")
        repaired_cost = input("Jaka jest cena naprawy?: ")
        repaired_list.update({repaired_thing:repaired_cost})
        next_repaired = input("Czy chcesz dodać kolejną naprawę?: y/n ").lower()
        if next_repaired == "y":
            continue
        elif next_repaired != "n":
            continue
        else:
            break
    return repaired_list

def repair_fast_func():
    """inputing recommendation for short term repair"""
    repair_fast = input("Co należy zrobić jak najszybciej, w pierwszej kolejności?: ")
    recommendation = f"Pilne naprawy które należy wykonać jak najszybciej: {repair_fast}"
    return recommendation

def repair_long_func():
    """Inputing recommendatiob for long term repair"""
    repair_in_time = input("Co należy naprawić przy następnym przeglądzie?: ")
    recommendation = f"Przed następnym przeglądem należy naprawić: {repair_in_time}"
    return recommendation

def comment_func():
    """Inputing lasts commments"""
    comments = input("Komentarz, dodatkowe informacje: ")
    last_comment = f"Komentarz, dodatkowe informacje: {comments}"
    return last_comment



if __name__ == "__main__":
    main()